"""Chat attachment upload, indexing, and retrieval helpers."""

from __future__ import annotations

import base64
import binascii
import hashlib
import html
import io
import json
import mimetypes
import os
import re
import shutil
import subprocess
import tempfile
import time
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field

from clawagents.desktop_stores.app_paths import uploads_dir
from clawagents.gateway.desktop_router import require_auth
from clawagents.utils.atomic_write import atomic_write_bytes, atomic_write_text


router = APIRouter(tags=["attachments"], dependencies=[require_auth()])

MAX_UPLOAD_BYTES = 15 * 1024 * 1024
MAX_PREVIEW_CHARS = 24_000
MAX_INDEX_CHARS = 250_000
CHUNK_CHARS = 4_000
CHUNK_OVERLAP = 400
DEFAULT_TOMBSTONE_RETENTION_DAYS = 30
ALLOWED_EXTS = {
    ".txt", ".md", ".csv", ".tsv", ".json", ".xml", ".log",
    ".pdf", ".docx", ".xlsx", ".pptx",
    ".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg",
}
TEXT_EXTS = {".txt", ".md", ".csv", ".tsv", ".json", ".xml", ".log"}
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg"}
OFFICE_EXTS = {".docx", ".xlsx", ".pptx"}


class AttachmentUploadBody(BaseModel):
    filename: str
    mime_type: str | None = None
    data_base64: str


class AttachmentRecord(BaseModel):
    id: str
    filename: str
    mime_type: str
    size: int
    path: str
    kind: str
    text_preview: str
    text_truncated: bool
    checksum: str
    chunks_count: int
    warnings: list[str] = Field(default_factory=list)
    created_at: float
    chunks_path: str
    deleted_at: float | None = None
    deduped: bool = False


class AttachmentUploadResponse(AttachmentRecord):
    pass


class AttachmentSearchBody(BaseModel):
    query: str = ""
    attachment_ids: list[str] | None = None
    limit: int = 5


class AttachmentChunk(BaseModel):
    attachment_id: str
    filename: str
    chunk_index: int
    score: int
    text: str


class AttachmentSearchResponse(BaseModel):
    chunks: list[AttachmentChunk]


def _safe_segment(value: str, fallback: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", value).strip(".-")
    return cleaned[:120] or fallback


def _chat_dir(chat_id: str) -> Path:
    return uploads_dir() / _safe_segment(chat_id, "chat")


def _manifest_path(chat_id: str) -> Path:
    return _chat_dir(chat_id) / "manifest.json"


def _read_manifest(chat_id: str) -> list[AttachmentRecord]:
    path = _manifest_path(chat_id)
    if not path.exists():
        return []
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    records: list[AttachmentRecord] = []
    for item in raw if isinstance(raw, list) else []:
        try:
            records.append(AttachmentRecord.model_validate(item))
        except Exception:  # noqa: BLE001
            continue
    return records


def _write_manifest(chat_id: str, records: list[AttachmentRecord]) -> None:
    try:
        retention_days = float(os.getenv("CLAWAGENTS_ATTACHMENT_TOMBSTONE_DAYS", str(DEFAULT_TOMBSTONE_RETENTION_DAYS)))
    except ValueError:
        retention_days = DEFAULT_TOMBSTONE_RETENTION_DAYS
    cutoff = time.time() - max(0, retention_days) * 24 * 60 * 60
    retained = [
        record for record in records
        if record.deleted_at is None or record.deleted_at >= cutoff
    ]
    payload = [record.model_dump() for record in retained]
    atomic_write_text(_manifest_path(chat_id), json.dumps(payload, indent=2, sort_keys=True))


def _decode_payload(data_base64: str) -> bytes:
    try:
        raw = base64.b64decode(data_base64, validate=True)
    except (binascii.Error, ValueError):
        raise HTTPException(status_code=400, detail="data_base64 is not valid base64")
    if len(raw) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail=f"attachment too large ({len(raw)} bytes)")
    return raw


def _truncate(text: str) -> tuple[str, bool]:
    if len(text) <= MAX_PREVIEW_CHARS:
        return text, False
    return text[:MAX_PREVIEW_CHARS] + "\n...", True


def _xml_text(raw: bytes) -> str:
    try:
        root = ElementTree.fromstring(raw)
    except ElementTree.ParseError:
        text = raw.decode("utf-8", errors="ignore")
        text = re.sub(r"<[^>]+>", "\n", text)
        parts = [html.unescape(part).strip() for part in text.splitlines()]
        return "\n".join(part for part in parts if part)
    parts = [node.strip() for node in root.itertext() if node and node.strip()]
    return "\n".join(parts)


def _zip_xml_text(raw: bytes, names: list[str]) -> str:
    out: list[str] = []
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        members = zf.namelist()[:300]
        for name in names:
            if name.endswith(".xml") and name in members:
                matches = [name]
            else:
                prefix = name if name.endswith("/") else name
                matches = [n for n in members if n.startswith(prefix) and n.endswith(".xml")]
            for match in matches[:120]:
                info = zf.getinfo(match)
                if info.file_size > 5 * 1024 * 1024:
                    continue
                text = _xml_text(zf.read(match))
                if text:
                    out.append(text)
    return "\n\n".join(out)


def _xlsx_fallback_text(raw: bytes) -> str:
    out: list[str] = []
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        names = zf.namelist()
        shared: list[str] = []
        if "xl/sharedStrings.xml" in names:
            shared = _xml_text(zf.read("xl/sharedStrings.xml")).splitlines()
        sheets = sorted(n for n in names if n.startswith("xl/worksheets/") and n.endswith(".xml"))
        for sheet in sheets[:20]:
            try:
                root = ElementTree.fromstring(zf.read(sheet))
            except ElementTree.ParseError:
                continue
            out.append(f"Sheet: {Path(sheet).stem}")
            for row in [node for node in root.iter() if node.tag.endswith("row")][:1000]:
                cells: list[str] = []
                for cell in [node for node in row if node.tag.endswith("c")][:80]:
                    value = next((child.text or "" for child in cell if child.tag.endswith("v")), "")
                    if cell.attrib.get("t") == "s" and value.isdigit():
                        idx = int(value)
                        value = shared[idx] if idx < len(shared) else value
                    cells.append(value)
                if any(cells):
                    out.append("\t".join(cells))
    return "\n".join(out)


def _extract_xlsx_text(raw: bytes) -> str:
    try:
        from openpyxl import load_workbook  # type: ignore

        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        out: list[str] = []
        for sheet in workbook.worksheets[:20]:
            out.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(max_row=1000, max_col=80, values_only=True):
                values = ["" if value is None else str(value) for value in row]
                if any(values):
                    out.append("\t".join(values).rstrip())
        return "\n".join(out)
    except Exception:  # noqa: BLE001
        return _xlsx_fallback_text(raw)


def _run_text_command(args: list[str], *, stdin: bytes | None = None, timeout: int = 8) -> str:
    try:
        completed = subprocess.run(
            args,
            input=stdin,
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,
            timeout=timeout,
            check=False,
        )
    except (OSError, subprocess.TimeoutExpired):
        return ""
    if completed.returncode != 0:
        return ""
    return completed.stdout.decode("utf-8", errors="replace").strip()


def _extract_pdf_text(raw: bytes) -> str:
    for module_name in ("pypdf", "PyPDF2"):
        try:
            module = __import__(module_name)
            reader = module.PdfReader(io.BytesIO(raw))
            text = "\n".join((page.extract_text() or "") for page in reader.pages[:50]).strip()
            if text:
                return text
        except Exception:  # noqa: BLE001
            pass

    if shutil.which("pdftotext"):
        with tempfile.NamedTemporaryFile(suffix=".pdf") as f:
            f.write(raw)
            f.flush()
            text = _run_text_command(["pdftotext", "-layout", f.name, "-"], timeout=10)
            if text:
                return text

    decoded = raw.decode("latin-1", errors="ignore")
    chunks = re.findall(r"\(([^()]{2,400})\)", decoded)
    cleaned = [html.unescape(c.replace(r"\)", ")").replace(r"\(", "(")).strip() for c in chunks]
    text = "\n".join(c for c in cleaned if c)
    if text:
        return text

    return _ocr_pdf(raw)


def _ocr_pdf(raw: bytes) -> str:
    if not (shutil.which("pdftoppm") and shutil.which("tesseract")):
        return ""
    with tempfile.TemporaryDirectory() as tmp:
        pdf = Path(tmp) / "input.pdf"
        pdf.write_bytes(raw)
        prefix = Path(tmp) / "page"
        try:
            subprocess.run(
                ["pdftoppm", "-png", "-f", "1", "-l", "3", str(pdf), str(prefix)],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                timeout=15,
                check=False,
            )
        except (OSError, subprocess.TimeoutExpired):
            return ""
        out: list[str] = []
        for image in sorted(Path(tmp).glob("page-*.png"))[:3]:
            text = _run_text_command(["tesseract", str(image), "stdout", "--psm", "6"], timeout=12)
            if text:
                out.append(text)
        return "\n\n".join(out)


def _ocr_image(raw: bytes, suffix: str) -> str:
    if suffix == ".svg":
        return _xml_text(raw) or raw.decode("utf-8", errors="ignore")
    if not shutil.which("tesseract"):
        return ""
    with tempfile.NamedTemporaryFile(suffix=suffix) as f:
        f.write(raw)
        f.flush()
        return _run_text_command(["tesseract", f.name, "stdout", "--psm", "6"], timeout=12)


def _preview(text: str, fallback: str) -> tuple[str, bool]:
    preview, truncated = _truncate(text)
    return preview or fallback, truncated


def _extract_text(filename: str, raw: bytes) -> tuple[str, str, str, bool]:
    ext = Path(filename).suffix.lower()
    if ext in TEXT_EXTS:
        text = raw.decode("utf-8", errors="replace")
        preview, truncated = _preview(text, "")
        return "text", text, preview, truncated
    if ext == ".docx":
        text = _zip_xml_text(raw, ["word/document.xml", "word/header", "word/footer"])
        preview, truncated = _preview(text, "(DOCX uploaded; no extractable text found)")
        return "document", text, preview, truncated
    if ext == ".pptx":
        text = _zip_xml_text(raw, ["ppt/slides/", "ppt/notesSlides/"])
        preview, truncated = _preview(text, "(PPTX uploaded; no extractable text found)")
        return "presentation", text, preview, truncated
    if ext == ".xlsx":
        text = _extract_xlsx_text(raw)
        preview, truncated = _preview(text, "(XLSX uploaded; no extractable text found)")
        return "spreadsheet", text, preview, truncated
    if ext == ".pdf":
        text = _extract_pdf_text(raw)
        preview, truncated = _preview(text, "(PDF uploaded; text extraction may require OCR or a PDF parser)")
        return "pdf", text, preview, truncated
    if ext in IMAGE_EXTS:
        text = _ocr_image(raw, ext)
        if text.strip():
            preview, truncated = _preview(text, "")
            return "image", text, preview, truncated
        fallback = "(image uploaded; use the stored path or Markdown image reference for visual analysis)"
        return "image", "", fallback, False
    return "binary", "", "(uploaded file has no text extractor)", False


def _sniff_mime(raw: bytes) -> str:
    head = raw[:512].lstrip()
    if raw.startswith(b"%PDF-"):
        return "application/pdf"
    if raw.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image/png"
    if raw.startswith(b"\xff\xd8\xff"):
        return "image/jpeg"
    if raw.startswith((b"GIF87a", b"GIF89a")):
        return "image/gif"
    if raw.startswith(b"BM"):
        return "image/bmp"
    if raw.startswith(b"RIFF") and raw[8:12] == b"WEBP":
        return "image/webp"
    if head.startswith((b"<svg", b"<?xml")) and b"<svg" in head[:300].lower():
        return "image/svg+xml"
    if raw.startswith(b"PK\x03\x04"):
        return "application/zip"
    if b"\x00" not in raw[:4096]:
        return "text/plain"
    return "application/octet-stream"


def _validate_type(filename: str, declared_mime: str | None, raw: bytes) -> tuple[str, list[str]]:
    ext = Path(filename).suffix.lower()
    sniffed = _sniff_mime(raw)
    expected = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    warnings: list[str] = []

    if ext in IMAGE_EXTS and not sniffed.startswith("image/"):
        raise HTTPException(status_code=415, detail=f"{ext} does not match uploaded bytes ({sniffed})")
    if ext == ".pdf" and sniffed != "application/pdf":
        raise HTTPException(status_code=415, detail=f"{ext} does not match uploaded bytes ({sniffed})")
    if ext in OFFICE_EXTS and sniffed != "application/zip":
        raise HTTPException(status_code=415, detail=f"{ext} does not match uploaded bytes ({sniffed})")
    if ext in TEXT_EXTS and sniffed == "application/octet-stream":
        raise HTTPException(status_code=415, detail=f"{ext} appears to be binary data")

    if declared_mime and declared_mime != "application/octet-stream":
        declared_family = declared_mime.split("/", 1)[0]
        expected_family = expected.split("/", 1)[0]
        if declared_family != expected_family and ext not in OFFICE_EXTS:
            warnings.append(f"declared MIME {declared_mime} differs from expected {expected}")
    return declared_mime or expected, warnings


def _chunks_for_text(text: str) -> list[dict[str, Any]]:
    source = text[:MAX_INDEX_CHARS]
    if not source:
        return []
    chunks: list[dict[str, Any]] = []
    start = 0
    index = 0
    while start < len(source):
        end = min(len(source), start + CHUNK_CHARS)
        chunk = source[start:end].strip()
        if chunk:
            chunks.append({"index": index, "text": chunk})
            index += 1
        if end >= len(source):
            break
        start = max(end - CHUNK_OVERLAP, start + 1)
    return chunks


def _write_chunks(path: Path, chunks: list[dict[str, Any]]) -> None:
    atomic_write_text(path, json.dumps(chunks, ensure_ascii=False))


def _load_chunks(record: AttachmentRecord) -> list[dict[str, Any]]:
    try:
        raw = json.loads(Path(record.chunks_path).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    return raw if isinstance(raw, list) else []


def _score_chunk(query_terms: set[str], text: str, filename: str) -> int:
    haystack = f"{filename}\n{text}".lower()
    return sum(haystack.count(term) for term in query_terms)


def _active_records(chat_id: str, attachment_ids: set[str] | None = None) -> list[AttachmentRecord]:
    records = [record for record in _read_manifest(chat_id) if record.deleted_at is None]
    if attachment_ids is not None:
        records = [record for record in records if record.id in attachment_ids]
    return records


def _response(record: AttachmentRecord, *, deduped: bool = False) -> AttachmentUploadResponse:
    data = record.model_dump()
    data["deduped"] = deduped
    return AttachmentUploadResponse.model_validate(data)


@router.post("/chats/{chat_id}/attachments", status_code=201)
def upload_chat_attachment(chat_id: str, body: AttachmentUploadBody) -> AttachmentUploadResponse:
    safe_name = _safe_segment(Path(body.filename).name, "attachment")
    ext = Path(safe_name).suffix.lower()
    if ext not in ALLOWED_EXTS:
        raise HTTPException(status_code=415, detail=f"unsupported attachment type: {ext or '(none)'}")

    raw = _decode_payload(body.data_base64)
    mime_type, warnings = _validate_type(safe_name, body.mime_type, raw)
    checksum = f"sha256:{hashlib.sha256(raw).hexdigest()}"
    records = _read_manifest(chat_id)
    for record in records:
        if record.deleted_at is None and record.checksum == checksum:
            return _response(record, deduped=True)

    attachment_id = f"{int(time.time() * 1000)}-{checksum[7:15]}-{safe_name}"
    target_dir = _chat_dir(chat_id)
    target_path = target_dir / attachment_id
    chunks_path = target_dir / f"{attachment_id}.chunks.json"
    atomic_write_bytes(target_path, raw)
    kind, full_text, preview, truncated = _extract_text(safe_name, raw)
    chunks = _chunks_for_text(full_text or preview)
    _write_chunks(chunks_path, chunks)

    if kind == "image":
        warnings.append(f"vision reference: ![{safe_name}]({target_path})")

    record = AttachmentRecord(
        id=attachment_id,
        filename=safe_name,
        mime_type=mime_type,
        size=len(raw),
        path=str(target_path),
        kind=kind,
        text_preview=preview,
        text_truncated=truncated,
        checksum=checksum,
        chunks_count=len(chunks),
        warnings=warnings,
        created_at=time.time(),
        chunks_path=str(chunks_path),
    )
    records.append(record)
    _write_manifest(chat_id, records)
    return _response(record)


@router.get("/chats/{chat_id}/attachments")
def list_chat_attachments(chat_id: str) -> list[AttachmentRecord]:
    return _active_records(chat_id)


@router.get("/chats/{chat_id}/attachments/{attachment_id}/download")
def download_chat_attachment(chat_id: str, attachment_id: str) -> FileResponse:
    record = next((item for item in _active_records(chat_id) if item.id == attachment_id), None)
    if record is None or not Path(record.path).exists():
        raise HTTPException(status_code=404, detail="attachment not found")
    return FileResponse(record.path, media_type=record.mime_type, filename=record.filename)


@router.delete("/chats/{chat_id}/attachments/{attachment_id}")
def delete_chat_attachment(chat_id: str, attachment_id: str) -> dict:
    records = _read_manifest(chat_id)
    for index, record in enumerate(records):
        if record.id != attachment_id or record.deleted_at is not None:
            continue
        for path in (Path(record.path), Path(record.chunks_path)):
            try:
                path.unlink()
            except FileNotFoundError:
                pass
            except OSError:
                pass
        records[index] = record.model_copy(update={"deleted_at": time.time()})
        _write_manifest(chat_id, records)
        return {"ok": True}
    raise HTTPException(status_code=404, detail="attachment not found")


@router.post("/chats/{chat_id}/attachments/search")
def search_chat_attachments(chat_id: str, body: AttachmentSearchBody) -> AttachmentSearchResponse:
    query_terms = {term for term in re.findall(r"[A-Za-z0-9_]{2,}", body.query.lower())}
    selected = set(body.attachment_ids) if body.attachment_ids else None
    hits: list[AttachmentChunk] = []
    for record in _active_records(chat_id, selected):
        for chunk in _load_chunks(record):
            text = str(chunk.get("text") or "")
            score = _score_chunk(query_terms, text, record.filename) if query_terms else 1
            if score <= 0:
                continue
            hits.append(AttachmentChunk(
                attachment_id=record.id,
                filename=record.filename,
                chunk_index=int(chunk.get("index") or 0),
                score=score,
                text=text,
            ))
    hits.sort(key=lambda item: item.score, reverse=True)
    return AttachmentSearchResponse(chunks=hits[: max(1, min(body.limit, 20))])


def build_attachment_context(chat_id: str, query: str, attachment_ids: list[str] | None) -> tuple[str, list[dict[str, Any]]]:
    """Return prompt context plus metadata for the visible user message."""
    selected = set(attachment_ids) if attachment_ids else None
    records = _active_records(chat_id, selected)
    if not records:
        return "", []

    visible = [record.model_dump(exclude={"chunks_path"}) for record in records if selected is not None]
    search = search_chat_attachments(
        chat_id,
        AttachmentSearchBody(query=query, attachment_ids=attachment_ids, limit=8 if selected else 4),
    )
    parts: list[str] = []
    if selected is not None:
        parts.append("Attached files for this turn:")
        for record in records:
            image_ref = f"\n- vision reference: ![{record.filename}]({record.path})" if record.kind == "image" else ""
            parts.append(
                f"### {record.filename}\n"
                f"- id: {record.id}\n- kind: {record.kind}\n- MIME: {record.mime_type}\n"
                f"- size: {record.size} bytes\n- path: {record.path}{image_ref}\n"
                f"- warnings: {', '.join(record.warnings) if record.warnings else 'none'}"
            )
    if search.chunks:
        parts.append("Relevant uploaded-file chunks:")
        for chunk in search.chunks:
            parts.append(f"### {chunk.filename} chunk {chunk.chunk_index}\n```text\n{chunk.text}\n```")
    return "\n\n".join(parts), visible


_MAX_IMAGES_PER_TURN = 8
_MAX_FILES_PER_TURN = 4
_MAX_B64_BYTES = 14 * 1024 * 1024


def build_invoke_media(
    chat_id: str,
    attachment_ids: list[str] | None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Load uploaded attachments as native ``images`` / ``files`` for ``agent.invoke``.

    Images go to vision; PDF/DOCX (and other docs) go through the library's
    document path. Returns (images, files) as base64 payloads.
    """
    import base64

    if not attachment_ids:
        return [], []
    selected = set(attachment_ids)
    records = _active_records(chat_id, selected)
    images: list[dict[str, Any]] = []
    files: list[dict[str, Any]] = []
    for record in records:
        try:
            raw = Path(record.path).read_bytes()
        except OSError:
            continue
        b64 = base64.b64encode(raw).decode("ascii")
        if len(b64) > _MAX_B64_BYTES:
            continue
        mime = (record.mime_type or "application/octet-stream").strip()
        if record.kind == "image" or mime.startswith("image/"):
            if len(images) >= _MAX_IMAGES_PER_TURN:
                continue
            images.append({"data": b64, "media_type": mime or "image/png"})
        else:
            if len(files) >= _MAX_FILES_PER_TURN:
                continue
            files.append({
                "data": b64,
                "media_type": mime or "application/pdf",
                "name": record.filename or "attachment",
            })
    return images, files
